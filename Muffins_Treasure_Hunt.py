import os
import re
import logging
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PyPDF2 import PdfReader  # Removed PdfReadWarning import
from docx import Document
import time
import pytesseract
from PIL import Image
import warnings
import sys

# ---------------------------
# Configuration Settings
# ---------------------------

# Allowed file extensions (in lowercase)
ALLOWED_EXTENSIONS = [
    '', '.txt', '.docx', '.pdf', '.png', '.jpg', '.jpeg',
    '.zip', '.one', '.json', '.csv', '.xlsx'
]

# Keywords to search for with corresponding emojis
KEYWORDS_EMOJIS = {
    "crypto": "🔒",
    "wallet": "👛",
    "bitcoin": "₿",
    "ethereum": "Ξ",
    "doge": "🐕",
    "key": "🗝️",
    "phrase": "📝",
    "secret": "🤫",
    "password": "🔑",
    "passphrase": "🔑",
    "xpub": "📈",
    "token": "🎫",
    "backup": "💾",
    "seed": "🌱",
    "private": "🔐",
    "credentials": "🛂",
    "blockchain": "⛓️",
    "coins": "💰",
    "hash": "🔢",
    "wallet.dat": "📂",
    "mnemonic": "🧠",
    "recovery": "🔄",
    "restore": "🛠️",
    "seed phrase": "🌱📝",
    "secret phrase": "🤫📝",
    "metamask": "🐒",
    "phantom": "👻",
    "keystore": "🗄️",
    "ledger": "📚",
    "trezor": "🔐",
    "cold storage": "❄️📦",
    "private_key": "🔑",
    "xprv": "📈🔑",
    "encrypted": "🔒",
    "kdfparams": "⚙️",
    "cipher": "🔡",
    "ciphertext": "🔠",
    "btc": "₿",
    "eth": "Ξ",
    "exodus": "🗺️",
    "trustwallet": "🤝👛",
    "binance": "💹",
    "kraken": "🦑",
    "seed_phrase": "🌱📝",
    "recovery_seed": "🔄🌱",
    "backup_phrase": "💾📝"
}

# Excluded directories based on absolute paths (Windows and Mac)
EXCLUDED_DIRS = [
    # Windows System Directories
    r"C:\Windows",
    r"C:\adobeTemp",
    r"C:\Users\Public",
    r"C:\$Recycle.Bin",
    # Mac System Directories
    "/System",
    "/Library",
    "/private",
    "/usr"
]

# Excluded directory names (exact matches)
EXCLUDED_DIR_NAMES = [
    'AppData',
    '.vscode',
    'Quarantine_Unwanted_Files',
    'ProgramData',
    'mysy64',
    'cygwin64',
    'Oculus',
    'steam',
    'Unity'
]

# Excluded directory name patterns (regex patterns for partial matches)
EXCLUDED_DIR_PATTERNS = [
    r'^python\d*$',            # Matches 'python', 'python3', 'python311', etc.
    r'^adobe.*$',              # Matches 'Adobe', 'AdobeSuite', etc.
    r'^unity.*$',              # Matches 'Unity', 'Unity2021', etc.
    r'^vscode.*$',             # Matches 'vscode', 'vscode-insiders', etc.
    r'^steam.*$',              # Matches 'steam', 'steamgames', etc.
    r'^game_libs$',            # Example pattern for game libraries
    r'^cygwin.*$',             # Matches 'cygwin64', 'cygwin32', etc.
    r'^quarantine.*$'          # Matches 'quarantine', 'quarantine_unwanted_files', etc.
]

# Path for the results
HOME_DIR = Path.home()
DESKTOP_PATH = HOME_DIR / "Desktop" / "Muffins_Treasure_Hunt_Results"
CONSOLE_LOG_FILE = DESKTOP_PATH / "Muffins_Treasure_Hunt_Console_Log.txt"
ERROR_LOG_FILE = DESKTOP_PATH / "Muffins_Treasure_Hunt_Errors.txt"

# Ensure the results directory exists
DESKTOP_PATH.mkdir(parents=True, exist_ok=True)

# Setup logging
logging.basicConfig(
    level=logging.INFO,  # Set to DEBUG for more detailed logs
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(CONSOLE_LOG_FILE, mode='w', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

# Suppress PyPDF2 warnings
warnings.filterwarnings("ignore", category=UserWarning)  # Changed category to UserWarning
logging.getLogger("PyPDF2").setLevel(logging.ERROR)

# ---------------------------
# Function to Print Intro Branding
# ---------------------------

def print_intro():
    # Define the width for centering
    WIDTH = 60
    
    # Centered headers
    header_line = "=" * WIDTH
    title = "🐶 Muffin's Treasure Hunt 🐾"
    subtitle = "🏆 Sniffing Out Crypto Treasures! 🏆"
    
    print(header_line)
    print(title.center(WIDTH))
    print(subtitle.center(WIDTH))
    print(header_line)
    print()
    
    # Introduction message with wrapped lines
    intro_lines = [
        "Muffin dog is here to help you uncover hidden crypto treasures on your computer.",
        "This tool searches through your drives for crypto wallets, keys, recovery phrases, and other sensitive information.",
        ""
    ]
    
    for line in intro_lines:
        print(line)
    
    # Features with bullets
    print("🔍 What Does Muffin's Treasure Hunt Do?")
    features = [
        "- 🦴 Searches Files: Scans files like .txt, .docx, .pdf, and more for crypto-related keywords.",
        "- 🐕 Performs OCR: Extracts text from image files to find hidden information.",
        "- 📂 Excludes System Directories: Skips directories and folders that are unlikely to contain relevant data to reduce noise.",
        "- 💼 Generates Reports: Outputs results to both a text file and an Excel spreadsheet for easy review."
    ]
    
    for feature in features:
        print(feature)
    
    print()
    print(header_line)

# ---------------------------
# Function to Print Outro Branding
# ---------------------------

def print_outro(total_found, total_time_seconds):
    # Convert total_time_seconds to hours, minutes, and seconds
    hours, remainder = divmod(int(total_time_seconds), 3600)
    minutes, seconds = divmod(remainder, 60)
    time_formatted = f"{hours}h {minutes}m {seconds}s"

    outro_message = f"""
=============================================================
🎉 Muffin's Treasure Hunt Complete! 🎉
🏆 Total Treasures Found: {total_found}
⏰ Total Time Taken: {time_formatted}
📊 Results Exported: Check the Excel and Text files on your Desktop.
=============================================================
"""
    print(outro_message)

# ---------------------------
# Function to Load BIP39 Wordlist
# ---------------------------

def load_bip39_wordlist(wordlist_path):
    """
    Loads the BIP39 wordlist from a text file into a set for quick lookup.
    Each word should be on a separate line in the file.
    """
    try:
        with open(wordlist_path, 'r', encoding='utf-8') as f:
            words = set(word.strip().lower() for word in f if word.strip())
        logger.info(f"✅ Loaded BIP39 wordlist with {len(words)} words.")
        return words
    except FileNotFoundError:
        logger.error(f"❌ BIP39 wordlist file not found at: {wordlist_path}")
        return set()
    except Exception as e:
        logger.error(f"❌ Error loading BIP39 wordlist: {e}")
        return set()

# ---------------------------
# Function to Load Keywords
# ---------------------------

def load_keywords():
    """
    Returns a compiled regex pattern for the keywords.
    The pattern is case-insensitive and matches whole words.
    """
    escaped_keywords = [re.escape(keyword) for keyword in KEYWORDS_EMOJIS.keys()]
    pattern_string = r'\b(' + '|'.join(escaped_keywords) + r')\b'
    return re.compile(pattern_string, re.IGNORECASE)

KEYWORD_PATTERN = load_keywords()

# ---------------------------
# Function to Check Allowed Extensions
# ---------------------------

def is_allowed_file(file_path):
    """
    Checks if the file has an allowed extension or no extension.
    """
    return file_path.suffix.lower() in ALLOWED_EXTENSIONS

# ---------------------------
# Function to Exclude Directories
# ---------------------------

def should_exclude_dir(dir_path):
    """
    Determines whether a directory should be excluded based on:
    - Its absolute path
    - Its name (exact matches and pattern-based matches)
    """
    # Check absolute paths
    for excluded in EXCLUDED_DIRS:
        try:
            if dir_path.resolve() == Path(excluded).resolve():
                logger.debug(f"Excluding directory by absolute path: {dir_path}")
                return True
        except Exception as e:
            logger.debug(f"Error resolving path {dir_path}: {e}")
            continue

    # Check directory names for exact matches
    dir_name = dir_path.name.lower()
    for excluded_name in EXCLUDED_DIR_NAMES:
        if excluded_name.lower() == dir_name:
            logger.debug(f"Excluding directory by name: {dir_path}")
            return True

    # Check directory names against exclusion patterns
    for pattern in EXCLUDED_DIR_PATTERNS:
        if re.match(pattern, dir_name):
            logger.debug(f"Excluding directory by pattern: {dir_path}")
            return True

    return False

# ---------------------------
# Function to Search File Content for Keywords and Seed Phrases
# ---------------------------

def search_file_content(file_path, extension, bip39_words):
    """
    Searches for keywords and BIP39 seed phrases in the content of the file.
    Returns a tuple (keywords_found, seed_phrases_found).
    """
    keywords_found = set()
    seed_phrases_found = []

    try:
        if extension == '.txt':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
        elif extension == '.docx':
            doc = Document(file_path)
            content = "\n".join([para.text for para in doc.paragraphs])
        elif extension == '.pdf':
            reader = PdfReader(file_path)
            content = "\n".join([page.extract_text() or '' for page in reader.pages])
        elif extension == '.one':
            # OneNote files are proprietary; content extraction is non-trivial.
            # You might need a specialized library or skip content search for '.one' files.
            logger.debug(f"Skipping OneNote file content search: {file_path}")
            content = ""
        elif extension in ['.png', '.jpg', '.jpeg']:
            # Perform OCR on image files
            content = ocr_image(file_path)
        else:
            # For other file types, skip content search
            content = ""

        if content:
            # Search for keywords
            keyword_matches = KEYWORD_PATTERN.findall(content)
            if keyword_matches:
                keywords_found.update(match.lower() for match in keyword_matches)

            # Search for seed phrases
            seed_phrases_found = find_seed_phrases(content, bip39_words)

            # Search for Bitcoin and Ethereum addresses
            bitcoin_addresses = find_bitcoin_addresses(content)
            ethereum_addresses = find_ethereum_addresses(content)

            # Combine all found seed phrases and addresses
            if bitcoin_addresses:
                seed_phrases_found.extend(bitcoin_addresses)
            if ethereum_addresses:
                seed_phrases_found.extend(ethereum_addresses)

    except Exception as e:
        logger.debug(f"Error reading file {file_path}: {e}")

    return keywords_found, seed_phrases_found

def ocr_image(file_path):
    """
    Extracts text from an image file using OCR.
    Returns the extracted text as a string.
    """
    try:
        image = Image.open(file_path)
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        logger.debug(f"Error performing OCR on {file_path}: {e}")
        return ""

def find_seed_phrases(text, bip39_words):
    """
    Searches the text for BIP39 seed phrases.
    Returns a list of found seed phrases.
    """
    seed_phrases = []
    # Normalize text to lowercase
    text = text.lower()

    # Split text into words
    words = re.findall(r'\b\w+\b', text)

    # Define possible seed phrase lengths
    phrase_lengths = [12, 15, 18, 21, 24]

    # Slide through the words and check for valid seed phrases
    for length in phrase_lengths:
        for i in range(len(words) - length + 1):
            phrase = words[i:i + length]
            if all(word in bip39_words for word in phrase):
                seed_phrases.append(' '.join(phrase))

    return seed_phrases

def find_bitcoin_addresses(text):
    """
    Searches the text for Bitcoin addresses.
    Returns a list of found Bitcoin addresses.
    """
    bitcoin_addresses = set()
    # Regex patterns for different Bitcoin address formats
    patterns = [
        r'\b[13][a-km-zA-HJ-NP-Z1-9]{25,34}\b',                # P2PKH and P2SH
        r'\bbc1[a-z0-9]{11,71}\b'                            # Bech32
    ]

    for pattern in patterns:
        matches = re.findall(pattern, text)
        for match in matches:
            bitcoin_addresses.add(match)

    return list(bitcoin_addresses)

def find_ethereum_addresses(text):
    """
    Searches the text for Ethereum addresses.
    Returns a list of found Ethereum addresses.
    """
    ethereum_addresses = set()
    # Regex pattern for Ethereum addresses
    pattern = r'\b0x[a-fA-F0-9]{40}\b'
    matches = re.findall(pattern, text)
    for match in matches:
        ethereum_addresses.add(match)

    return list(ethereum_addresses)

# ---------------------------
# Function to Process a Single File
# ---------------------------

def process_file(file_path, bip39_words):
    """
    Processes a single file:
    - Checks for keywords in the filename
    - If applicable, checks for keywords, seed phrases, and crypto addresses in the file content
    Returns a dictionary with file details if matches are found, else None.
    """
    matched_keywords = set()
    matched_seed_phrases = []
    matched_bitcoin_addresses = []
    matched_ethereum_addresses = []
    file = Path(file_path)
    extension = file.suffix.lower()

    # Check filename for keywords
    if KEYWORD_PATTERN.search(file.name):
        matches = KEYWORD_PATTERN.findall(file.name)
        matched_keywords.update(match.lower() for match in matches)

    # Check content for keywords, seed phrases, and addresses if the file is text-based or image
    if extension in ['.txt', '.docx', '.pdf', '.one', '.png', '.jpg', '.jpeg']:
        keywords_in_content, seed_phrases_or_addresses = search_file_content(file, extension, bip39_words)
        if keywords_in_content:
            matched_keywords.update(keywords_in_content)
        for item in seed_phrases_or_addresses:
            if re.match(r'^0x[a-fA-F0-9]{40}$', item):
                matched_ethereum_addresses.append(item)
            elif re.match(r'^[13][a-km-zA-HJ-NP-Z1-9]{25,34}$', item) or re.match(r'^bc1[a-z0-9]{11,71}$', item):
                matched_bitcoin_addresses.append(item)
            else:
                matched_seed_phrases.append(item)

    if matched_keywords or matched_seed_phrases or matched_bitcoin_addresses or matched_ethereum_addresses:
        # Assign emojis to keywords found
        emojis = ''.join([KEYWORDS_EMOJIS.get(keyword, "") for keyword in matched_keywords])

        # Determine Top Folder (highest folder after the drive)
        path_parts = file.parent.resolve().parts
        if os.name == 'nt':
            # Remove the drive letter
            path_parts = path_parts[1:]
        else:
            # For Unix-like systems, first part is root '/'
            path_parts = path_parts[1:]

        top_folder = path_parts[0] if len(path_parts) > 0 else "Root"

        # Determine Bottom Folder (immediate folder containing the file)
        bottom_folder = file.parent.name if file.parent.name else "Root"

        # Compile all found seed phrases and addresses
        combined_seed_phrases = matched_seed_phrases + matched_bitcoin_addresses + matched_ethereum_addresses

        result = {
            "Drive": file.drive if os.name == 'nt' else "/",
            "Top Folder": top_folder,
            "Bottom Folder": bottom_folder,
            "Keyword Match": ", ".join(matched_keywords) if matched_keywords else "None",
            "Seed Phrases/Addresses Found": "; ".join(combined_seed_phrases) if combined_seed_phrases else "None",
            "File Extension": extension if extension else "No Extension",
            "File Name": file.name,
            "File Path": str(file.parent.resolve())
        }
        logger.info(f"{emojis} Found in: {file}")
        return result
    return None

# ---------------------------
# Function to Traverse Directories
# ---------------------------

def traverse_directories(root_dir, max_depth=5):
    """
    Traverses directories starting from root_dir, excluding specified directories.
    Limits the traversal to max_depth levels.
    Returns a list of file paths that match the criteria.
    """
    matched_files = []
    for dirpath, dirnames, filenames in os.walk(root_dir, topdown=True):
        current_dir = Path(dirpath)
        logger.debug(f"Scanning directory: {current_dir}")

        # Calculate current depth
        try:
            relative_path = current_dir.relative_to(root_dir)
            depth = len(relative_path.parts)
        except ValueError:
            # If current_dir is not a subdirectory of root_dir
            depth = 0

        if depth > max_depth:
            dirnames[:] = []  # Don't traverse deeper
            logger.debug(f"Max depth reached at {current_dir}. Skipping subdirectories.")
            continue

        # Modify dirnames in-place to skip excluded directories
        original_dirnames = list(dirnames)
        try:
            dirnames[:] = [d for d in dirnames if not should_exclude_dir(current_dir / d)]
            excluded_dirs_now = set(original_dirnames) - set(dirnames)
            if excluded_dirs_now:
                logger.debug(f"Excluded directories in {current_dir}: {', '.join(excluded_dirs_now)}")
        except Exception as e:
            logger.error(f"Error excluding directories in {current_dir}: {e}")
            continue

        for file in filenames:
            file_path = current_dir / file
            if is_allowed_file(file_path):
                matched_files.append(str(file_path))
    logger.info(f"Found {len(matched_files)} files to process in {root_dir}")
    return matched_files

# ---------------------------
# Function to Export Results
# ---------------------------

def export_results(found_items):
    """
    Exports the found items to a text file and an Excel spreadsheet.
    """
    text_file = DESKTOP_PATH / "Muffins_Treasure_Hunt_Path_Log.txt"
    excel_file = DESKTOP_PATH / "Muffins_Treasure_Hunt_Results.xlsx"

    # Export to Text File
    try:
        with open(text_file, "w", encoding="utf-8") as txt:
            txt.write("🔍 Muffin's Treasure Hunt Results\n")
            txt.write(f"🏆 Total treasures found: {len(found_items)}\n\n")
            for item in found_items:
                txt.write(f"Drive: {item['Drive']} | Top Folder: {item['Top Folder']} | Bottom Folder: {item['Bottom Folder']} | File: {item['File Name']} | Path: {item['File Path']}\n")
                if item["Seed Phrases/Addresses Found"] != "None":
                    txt.write(f"🔑 Seed Phrases/Addresses: {item['Seed Phrases/Addresses Found']}\n")
                txt.write("\n")
        logger.info(f"📄 Text File: {text_file}")
    except Exception as e:
        logger.error(f"Failed to write text file: {e}")

    # Export to Excel Spreadsheet
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Muffin's Results"

        headers = [
            "Drive", "Top Folder", "Bottom Folder",
            "Keyword Match", "Seed Phrases/Addresses Found",
            "File Extension", "File Name", "File Path"
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for row, item in enumerate(found_items, 2):
            sheet.cell(row=row, column=1, value=item["Drive"])
            sheet.cell(row=row, column=2, value=item["Top Folder"])
            sheet.cell(row=row, column=3, value=item["Bottom Folder"])
            sheet.cell(row=row, column=4, value=item["Keyword Match"])
            sheet.cell(row=row, column=5, value=item["Seed Phrases/Addresses Found"])
            sheet.cell(row=row, column=6, value=item["File Extension"])
            sheet.cell(row=row, column=7, value=item["File Name"])
            path_cell = sheet.cell(row=row, column=8, value=item["File Path"])
            path_cell.hyperlink = f"file:///{item['File Path']}"

        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 25

        workbook.save(excel_file)
        logger.info(f"📊 Spreadsheet: {excel_file}")
    except Exception as e:
        logger.error(f"Failed to write Excel file: {e}")

    logger.info(f"🏆 Total treasures found: {len(found_items)} 🐾")

# ---------------------------
# Function to Get Available Drives
# ---------------------------

def get_drives():
    """
    Retrieves all fixed drives on the system with branded messages.
    Adds an option to select 'ALL' drives.
    """
    drives = []
    if os.name == 'nt':
        import string
        import ctypes
        bitmask = ctypes.windll.kernel32.GetLogicalDrives()
        for letter in string.ascii_uppercase:
            if bitmask & 1:
                drive = f"{letter}:\\"
                if os.path.exists(drive):
                    drives.append(drive)
            bitmask >>= 1
    else:
        # For Unix-like systems, root is '/'
        drives.append('/')

    # Branded message about the drives being scanned
    logger.info(f"🔍 Preparing to scan the following drives: {', '.join(drives)}")
    return drives

# ---------------------------
# Main Function
# ---------------------------

def main():
    """
    Main function to execute the treasure hunt.
    """
    print_intro()  # Print the branded intro message
    start_time = time.time()
    logger.info("🔍 Starting Muffin's Treasure Hunt!")

    # Define root directories to scan (for production, scan all fixed drives)
    all_drives = get_drives()

    # Load BIP39 wordlist
    wordlist_path = Path(__file__).parent / "bip39_wordlist.txt"  # Assumes the wordlist is in the same directory
    bip39_words = load_bip39_wordlist(wordlist_path)

    if not bip39_words:
        logger.warning("BIP39 wordlist is empty or not loaded. Seed phrase detection will be skipped.")

    # Prompt user to input drive letters separated by spaces or type 'ALL'
    print("Available Drives:")
    for drive in all_drives:
        print(f"- {drive}")
    print("- ALL")  # Option to select all drives
    print("\nEnter the drive letters you want to scan, separated by spaces (e.g., C D E) or type ALL to scan all drives: ", end="", flush=True)

    user_input = sys.stdin.readline().strip().upper()
    if user_input == "ALL":
        selected_drives = all_drives
    else:
        selected_drives = [f"{letter.upper()}:\\"
                           for letter in user_input.split()
                           if f"{letter.upper()}:\\"
                           in all_drives]

    if not selected_drives:
        logger.warning("No valid drives selected. Exiting.")
        print("No valid drives selected. Exiting.")
        return

    print(f"\n🚀 Starting search on drives: {', '.join(selected_drives)}\n", flush=True)
    all_found_items = []

    # Use ThreadPoolExecutor for concurrent processing
    with ThreadPoolExecutor(max_workers=os.cpu_count() or 4) as executor:
        future_to_file = {}
        for root in selected_drives:
            logger.info(f"📂 Scanning directory: {root}")
            matched_files = traverse_directories(root, max_depth=5)
            logger.info(f"Matched {len(matched_files)} files in {root}")
            for file_path in matched_files:
                future = executor.submit(process_file, file_path, bip39_words)
                future_to_file[future] = file_path

        # Use tqdm for progress bar
        if future_to_file:
            for future in tqdm(as_completed(future_to_file), total=len(future_to_file), desc="Scanning files"):
                try:
                    result = future.result()
                    if result:
                        all_found_items.append(result)
                except Exception as e:
                    file_path = future_to_file[future]
                    logger.error(f"Error processing file {file_path}: {e}")
        else:
            logger.info("No files matched the criteria to process.")

    # Export the results
    export_results(all_found_items)

    end_time = time.time()
    total_time_seconds = end_time - start_time

    # Remove the extra logging line that was printing the total time taken before the outro
    # logger.info(f"\n⏰ Total time taken: {total_time_seconds:.2f} seconds")

    print_outro(len(all_found_items), total_time_seconds)  # Print the outro branding

# ---------------------------
# Entry Point
# ---------------------------

if __name__ == "__main__":
    try:
        # Specify the path to the Tesseract executable if not in PATH
        # Uncomment and update the following line if Tesseract is not in your PATH
        # pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

        main()
    except KeyboardInterrupt:
        logger.warning("\n🛑 Hunt interrupted by user. Exiting gracefully.")
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}")
